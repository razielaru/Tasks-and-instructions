import streamlit as st
import json
import os
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ──────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="מערך הרבנות - פקודות ומשימות",
    page_icon="✡️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

BRIGADES = ['בנימין', 'שומרון', 'אפרים', 'מנשה', 'יהודה', 'עציון']
DATA_FILE = os.path.join(os.path.dirname(__file__), "mishimot_data.json")

IDF_GREEN = "#4E5A37"
IDF_LIGHT = "#7B8A62"
GOLD = "#D4AF37"

DEFAULT_TASKS = [
    {"id": "g1",  "title": 'היערכות לציוד קריאה בבתי הכנסת',  "desc": 'יש לבדוק שיש מגילת אסתר מקלף בכל בית כנסת ומגילות מנייר. שימו את מגילת הקלף בארון הקודש.',                                           "urgent": False, "category": "פורים"},
    {"id": "g2",  "title": 'קוראי מגילה במוצבים',              "desc": 'לוודא ולעדכן שבכל מוצב פלוגתי יש קורא מגילה (שם + טלפון). חבירה לחב"ד המרחבי.',                                                  "urgent": True,  "category": "פורים"},
    {"id": "g3",  "title": 'פרסום לו"ז פורים בבסיסים',         "desc": 'לפרסם לו"ז לתענית אסתר וקריאות מגילה: ערב (צאת הצום ולאחריו בחמ"ל), ובוקר (שחרית ולאחר שחרית).',                               "urgent": False, "category": "פורים"},
    {"id": "g4",  "title": 'הנחיות הלכה למעשה',                "desc": 'לפרסם הנחיות חטיבתיות על פי "הלכה כסדרה" - איך עושים משלוח מנות ומתנות לאביונים.',                                              "urgent": False, "category": "פורים"},
    {"id": "g5",  "title": 'תיאום סעודות החג',                 "desc": 'לתאם מול רס"ר מטבח: סעודה מפסקת (פתיחת צום), שבירת הצום, וסעודת החג.',                                                          "urgent": False, "category": "פורים"},
    {"id": "g6",  "title": 'כשרות תרומות',                     "desc": 'לוודא מול רס"ר המחנה את תרומות משלוחי המנות המגיעות לחיילים - שהכל כשר ללא ספק.',                                               "urgent": False, "category": "פורים"},
    {"id": "g7",  "title": 'קפ"ק חטיבתי מורחב',               "desc": 'לוודא שנקבע קפ"ק בראשות סמח"ט, קל"ח, רס"ר ונציגים רלוונטיים (מנהל מטבח).',                                                   "urgent": True,  "category": "פסח"},
    {"id": "g8",  "title": 'קפ"קים גדודיים',                   "desc": 'לוודא קיום קפ"קים בכל הגדודים (מג"ד, סמג"ד, מ"פים) בתיאום עם רב החטיבה. המטרה: היכשרות, ניקיונות והשבתות.',                    "urgent": False, "category": "פסח"},
    {"id": "g9",  "title": 'זימון אנשי מילואים',               "desc": 'לוודא שלכל "אנשי הפסח" יצא צו מילואים מסודר. דגש על חיילים שבפטור ובסיפוח.',                                                   "urgent": False, "category": "פסח"},
    {"id": "g10", "title": 'איתור כוח אדם נוסף',               "desc": 'להמשיך לחפש ולאתר אנשי מילואים נוספים לטובת מאמץ ההכשרות במחנות ובמוצבים.',                                                    "urgent": False, "category": "פסח"},
]

# ─── STYLING ─────────────────────────────────────────────────────────────────
def inject_css():
    st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Heebo:wght@300;400;500;700;800&display=swap');

    html, body, [class*="css"] {{ font-family: 'Heebo', sans-serif !important; direction: rtl; }}
    .stApp {{ background-color: #f1f5f0; }}

    /* Header */
    .app-header {{
        background: linear-gradient(135deg, {IDF_GREEN} 0%, #2E3521 100%);
        color: white;
        padding: 20px 28px;
        border-radius: 16px;
        margin-bottom: 24px;
        display: flex;
        align-items: center;
        gap: 16px;
        box-shadow: 0 4px 20px rgba(46,53,33,0.3);
    }}
    .app-header h1 {{ margin: 0; font-size: 1.6rem; font-weight: 800; }}
    .app-header p  {{ margin: 0; font-size: 0.85rem; color: #c8d0b0; }}
    .gold-badge {{ background: {GOLD}; color: #2E3521; padding: 4px 12px; border-radius: 20px; font-weight: 800; font-size: 0.8rem; }}

    /* Progress */
    .progress-wrap {{ background: #e5e7eb; border-radius: 999px; height: 10px; margin: 8px 0; }}
    .progress-bar  {{ background: {GOLD}; height: 10px; border-radius: 999px; transition: width .4s; }}

    /* Stat cards */
    .stat-card {{ background: white; border-radius: 14px; padding: 16px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
    .stat-num  {{ font-size: 2.2rem; font-weight: 800; margin: 0; }}
    .stat-lbl  {{ font-size: 0.75rem; color: #6b7280; margin: 0; }}

    /* Task items */
    .task-card {{
        background: white;
        border-radius: 12px;
        padding: 14px 18px;
        margin-bottom: 6px;
        border-right: 4px solid {IDF_GREEN};
        box-shadow: 0 1px 4px rgba(0,0,0,0.05);
    }}
    .task-card.done {{ border-right-color: #9ca3af; opacity: 0.7; background: #f8fafc; }}
    .task-card.urgent {{ border-right-color: #ef4444; }}
    .task-title {{ font-weight: 700; font-size: 1rem; color: #1f2937; }}
    .task-title.done {{ text-decoration: line-through; color: #9ca3af; }}
    .task-desc  {{ font-size: 0.85rem; color: #6b7280; margin-top: 2px; }}
    .badge {{ display: inline-block; padding: 2px 8px; border-radius: 20px; font-size: 0.72rem; font-weight: 700; margin-right: 6px; }}
    .badge-urgent {{ background: #fee2e2; color: #b91c1c; }}
    .badge-cat    {{ background: rgba(78,90,55,0.12); color: {IDF_GREEN}; }}

    /* Section header */
    .section-header {{ display: flex; align-items: center; gap: 10px; margin: 20px 0 10px; }}
    .section-icon {{ background: {IDF_GREEN}; color: white; border-radius: 10px; width: 36px; height: 36px; display: flex; align-items: center; justify-content: center; font-size: 1.1rem; flex-shrink: 0; }}
    .section-title {{ font-size: 1.2rem; font-weight: 800; color: #1f2937; margin: 0; }}

    /* Brigade card (overview) */
    .brigade-card {{
        background: white;
        border-radius: 14px;
        padding: 16px;
        text-align: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.07);
        cursor: pointer;
        transition: box-shadow .2s;
    }}
    .brigade-card:hover {{ box-shadow: 0 4px 16px rgba(0,0,0,0.12); }}
    .brigade-pct {{ font-size: 2rem; font-weight: 800; margin: 4px 0; }}

    /* Login */
    .login-card {{
        background: white;
        border-radius: 20px;
        padding: 40px;
        max-width: 440px;
        margin: 60px auto;
        box-shadow: 0 8px 32px rgba(0,0,0,0.1);
        text-align: center;
    }}
    .login-header {{
        background: linear-gradient(135deg, {IDF_GREEN}, #2E3521);
        color: white;
        border-radius: 14px;
        padding: 24px;
        margin-bottom: 24px;
    }}

    /* Streamlit overrides */
    div[data-testid="stCheckbox"] label {{ font-size: 1rem !important; font-weight: 600 !important; }}
    .stButton>button {{ border-radius: 999px !important; font-family: 'Heebo', sans-serif !important; font-weight: 700 !important; }}
    div[data-testid="stTabs"] [data-baseweb="tab"] {{ font-family: 'Heebo', sans-serif !important; font-weight: 700; }}
זה    st.markdown("""
    <div class="login-card">
        <div class="login-header">
            <div style="font-size:2.5rem">✡️</div>
            <h2 style="margin:8px 0 4px;font-size:1.5rem;font-weight:800">מערך הרבנות</h2>
            <p style="margin:0;color:#c8d0b0;font-size:0.85rem">פקודות ומשימות - פורים ופסח</p>
        </div>
        <h3 style="margin-bottom:20px;color:#374151">מי אתה?</h3>
    </div>
    """, unsafe_allow_html=True)

    # Center the buttons
    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        if st.button("👑  רב האוגדה", use_container_width=True, type="primary", key="login_rav"):
            st.session_state.current_user = "רב האוגדה"
            st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        for i, brigade in enumerate(BRIGADES):
            col = c1 if i % 2 == 0 else c2
            with col:
                if st.button(f'חטמ"ר {brigade}', use_container_width=True, key=f"login_{brigade}"):
                    st.session_state.current_user = brigade
                    st.rerun()

        st.markdown("<p style='text-align:center;color:#9ca3af;font-size:0.75rem;margin-top:16px'>הבחירה נשמרת בסשן הנוכחי</p>", unsafe_allow_html=True)

# ─── TASK ITEM RENDERER ───────────────────────────────────────────────────────
def render_task_row(task, checked, key, can_check=True, can_delete=False, on_check=None, on_delete=None):
    done_cls    = "done"    if checked else ""
    urgent_cls  = "urgent"  if (task.get("urgent") and not checked) else ""
    urgent_badge = '<span class="badge badge-urgent">⚠ חובת עדכון</span>' if task.get("urgent") else ""
    cat_badge    = f'<span class="badge badge-cat">{task["category"]}</span>' if task.get("category") else ""

    col_cb, col_content, col_del = st.columns([0.5, 8, 0.8] if can_delete else [0.5, 9, 0.1])
    with col_cb:
        if can_check:
            val = st.checkbox("", value=checked, key=key, label_visibility="collapsed")
            if val != checked and on_check:
                on_check(val)
        else:
            st.markdown(f"<div style='margin-top:8px;color:{'#4E5A37' if checked else '#d1d5db'};font-size:1.2rem'>{'✓' if checked else '○'}</div>", unsafe_allow_html=True)

    with col_content:
        st.markdown(f"""
        <div class="task-card {done_cls} {urgent_cls}" style="margin-bottom:0">
            <div class="task-title {done_cls}">{task['title']} {urgent_badge} {cat_badge}</div>
            {f'<div class="task-desc">{task["desc"]}</div>' if task.get("desc") else ""}
        </div>
        """, unsafe_allow_html=True)

    if can_delete:
        with col_del:
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            if st.button("🗑", key=f"del_{key}", help="מחק משימה"):
                if on_delete:
                    on_delete()

# ─── BRIGADE RABBI VIEW ───────────────────────────────────────────────────────
def show_brigade_view(brigade, data):
    total, done = completion_stats(data, brigade)
    left = total - done
    pct  = round((done / total) * 100) if total else 0

    # Stats
    c1, c2, c3 = st.columns(3)
    c1.markdown(f'<div class="stat-card"><p class="stat-num" style="color:{IDF_GREEN}">{total}</p><p class="stat-lbl">סה"כ משימות</p></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="stat-card"><p class="stat-num" style="color:#16a34a">{done}</p><p class="stat-lbl">הושלמו</p></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="stat-card"><p class="stat-num" style="color:#f97316">{left}</p><p class="stat-lbl">נותרו</p></div>', unsafe_allow_html=True)

    st.markdown(f"""
    <div class="progress-wrap">
        <div class="progress-bar" style="width:{pct}%"></div>
    </div>
    <p style="text-align:center;color:#6b7280;font-size:0.8rem">{pct}% הושלם</p>
    """, unsafe_allow_html=True)

    tab_global, tab_local, tab_schedule = st.tabs(["📡 משימות אוגדתיות", "📋 המשימות שלי", "📅 לו\"ז שבועי"])

    # ── Global tasks ──
    with tab_global:
        comp = data["completions"].setdefault(brigade, {})
        if not data["global_tasks"]:
            st.info("אין עדיין משימות אוגדתיות")
        for task in data["global_tasks"]:
            ck_key = f"g_{task['id']}"
            checked = bool(comp.get(ck_key, False))
            def make_check_fn(tid, ck):
                def fn(val):
                    data["completions"][brigade][ck] = val
                    save_data(data)
                    st.rerun()
                return fn
            render_task_row(task, checked, f"brigade_{brigade}_{ck_key}", on_check=make_check_fn(task["id"], ck_key))

    # ── Local tasks ──
    with tab_local:
        local_tasks = data["brigade_tasks"].get(brigade, [])
        comp = data["completions"].setdefault(brigade, {})

        if not local_tasks:
            st.info("אין עדיין משימות אישיות. הוסף משימה למטה.")

        for task in local_tasks:
            ck_key = f"l_{task['id']}"
            checked = bool(comp.get(ck_key, False))
            def make_check_fn(tid, ck):
                def fn(val):
                    data["completions"][brigade][ck] = val
                    save_data(data)
                    st.rerun()
                return fn
            def make_del_fn(tid):
                def fn():
                    data["brigade_tasks"][brigade] = [t for t in data["brigade_tasks"][brigade] if t["id"] != tid]
                    data["completions"][brigade].pop(f"l_{tid}", None)
                    save_data(data)
                    st.rerun()
                return fn
            render_task_row(task, checked, f"brigade_{brigade}_l_{task['id']}",
                            can_delete=True, on_check=make_check_fn(task["id"], ck_key),
                            on_delete=make_del_fn(task["id"]))

        st.markdown("---")
        st.markdown("**➕ הוסף משימה**")
        with st.form(f"add_local_{brigade}", clear_on_submit=True):
            title   = st.text_input("כותרת המשימה *", placeholder="כותרת קצרה וברורה")
            desc    = st.text_area("פירוט (אופציונלי)", placeholder="הוראות נוספות...", height=80)
            urgent  = st.checkbox("⚠ סמן כ\"חובת עדכון\"")
            submitted = st.form_submit_button("הוסף משימה", type="primary", use_container_width=True)
            if submitted and title.strip():
                data["brigade_tasks"].setdefault(brigade, []).append({
                    "id": gen_id(), "title": title.strip(), "desc": desc.strip(), "urgent": urgent, "category": ""
                })
                save_data(data)
                st.rerun()

    # ── Schedule ──
    with tab_schedule:
        show_schedule_tab(brigade, data)

# ─── SCHEDULE / CALENDAR TAB ─────────────────────────────────────────────────
EVENT_COLORS = {
    "אדום - חשוב":   {"bg": "#fef2f2", "border": "#ef4444", "dot": "#dc2626", "label": "🔴 חשוב"},
    "כתום - הכרחי":  {"bg": "#fff7ed", "border": "#f97316", "dot": "#ea580c", "label": "🟠 הכרחי"},
    "ירוק - כללי":   {"bg": "#f0fdf4", "border": "#22c55e", "dot": "#16a34a", "label": "🟢 כללי"},
}
HEB_DAYS = ['ראשון', 'שני', 'שלישי', 'רביעי', 'חמישי', 'שישי', 'שבת']

def _date_label(date_str):
    """Return Hebrew-friendly label like 'יום שני | 24/02'"""
    try:
        from datetime import date as dt_date
        d = dt_date.fromisoformat(date_str)
        dow = HEB_DAYS[d.weekday() % 7]   # Python: Mon=0; map to Sun=0 convention
        # Adjust: Python weekday 0=Mon; Sunday = weekday 6
        idf_dow = (d.weekday() + 1) % 7   # 0=Sun,1=Mon,...,6=Sat
        dow = HEB_DAYS[idf_dow]
        return f"יום {dow} | {d.strftime('%d/%m')}"
    except Exception:
        return date_str

def _build_event_html(ev, is_global=False):
    """Build event card HTML as a plain Python string (no nested f-string issues)."""
    color_key = ev.get("color", "ירוק - כללי")
    cs = EVENT_COLORS.get(color_key, EVENT_COLORS["ירוק - כללי"])
    bg      = cs["bg"]
    border  = cs["border"]
    dot_lbl = cs["label"]
    date_lbl = _date_label(ev.get("date", ""))
    time_part = (" &nbsp;🕐 " + ev["time"]) if ev.get("time") else ""
    desc_part = ('<p style="font-size:0.85rem;color:#4b5563;margin:4px 0 0">' + ev["desc"] + '</p>') if ev.get("desc") else ""
    global_badge = '<span style="background:#fef9c3;color:#92400e;padding:2px 8px;border-radius:999px;font-size:0.7rem;font-weight:700;margin-left:8px">★ אוגדתי</span>' if is_global else ""
    html = (
        '<div style="background:' + bg + ';border-radius:14px;padding:16px 20px;'
        'border-right:5px solid ' + border + ';box-shadow:0 2px 8px rgba(0,0,0,0.07);margin-bottom:10px">'
        '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px">'
        '<span style="font-size:0.8rem;font-weight:700;color:#374151">' + date_lbl + time_part + '</span>'
        '<span>' + dot_lbl + global_badge + '</span>'
        '</div>'
        '<div style="font-weight:800;font-size:1rem;color:#1f2937">' + ev["title"] + '</div>'
        + desc_part +
        '</div>'
    )
    return html


def show_calendar(events, key_prefix, can_delete=False, owner=None, data=None):
    """Render a 2-week calendar grid with events sorted by date."""
    from datetime import date as dt_date, timedelta
    today = dt_date.today()
    two_weeks = [today + timedelta(days=i) for i in range(14)]

    # Sort events by date
    def ev_sort_key(e):
        try: return dt_date.fromisoformat(e.get("date","9999-12-31"))
        except: return dt_date(9999,12,31)
    sorted_evs = sorted(events, key=ev_sort_key)

    # Group by date
    from collections import defaultdict
    by_date = defaultdict(list)
    for ev in sorted_evs:
        by_date[ev.get("date", "")].append(ev)

    # Build the 2-week grid header
    header_cells = ""
    for d in two_weeks:
        idf_dow = (d.weekday() + 1) % 7
        dow = HEB_DAYS[idf_dow]
        is_today = (d == today)
        bg_h = "#4E5A37" if is_today else "#f3f4f6"
        fg_h = "white" if is_today else "#374151"
        header_cells += (
            '<th style="text-align:center;padding:8px 4px;font-size:0.72rem;'
            'background:' + bg_h + ';color:' + fg_h + ';border-left:1px solid #e5e7eb;'
            'min-width:80px;font-weight:700">'
            + dow + '<br>' + d.strftime('%d/%m') +
            '</th>'
        )
    # Build event rows
    ev_cells = ""
    for d in two_weeks:
        date_key = d.isoformat()
        day_evs = by_date.get(date_key, [])
        cell_content = ""
        for ev in day_evs:
            cs = EVENT_COLORS.get(ev.get("color","ירוק - כללי"), EVENT_COLORS["ירוק - כללי"])
            is_global = ev.get("is_global", False)
            star = "★ " if is_global else ""
            cell_content += (
                '<div style="background:' + cs["bg"] + ';border-radius:8px;padding:6px 8px;'
                'margin-bottom:4px;border-right:3px solid ' + cs["border"] + ';'
                'font-size:0.72rem;font-weight:700;color:#1f2937;line-height:1.3">'
                + star + ev["title"] +
                (('<br><span style="font-weight:400;color:#6b7280">' + ev["time"] + '</span>') if ev.get("time") else "") +
                '</div>'
            )
        if not day_evs:
            cell_content = '<div style="height:24px"></div>'
        ev_cells += (
            '<td style="vertical-align:top;padding:6px 4px;border-left:1px solid #e5e7eb;'
            'border-top:1px solid #e5e7eb;min-width:80px">' + cell_content + '</td>'
        )

    grid_html = (
        '<div style="overflow-x:auto;background:white;border-radius:14px;'
        'box-shadow:0 2px 8px rgba(0,0,0,0.07);margin-bottom:16px">'
        '<table style="border-collapse:collapse;width:100%">'
        '<thead><tr>' + header_cells + '</tr></thead>'
        '<tbody><tr>' + ev_cells + '</tr></tbody>'
        '</table></div>'
    )
    st.markdown(grid_html, unsafe_allow_html=True)

    # List view below the grid with delete buttons
    if sorted_evs:
        st.markdown("**רשימת אירועים:**")
    for ev in sorted_evs:
        html = _build_event_html(ev, is_global=ev.get("is_global", False))
        col_card, col_del = st.columns([11, 1]) if can_delete else [st.container(), None]
        if can_delete:
            with col_card:
                st.markdown(html, unsafe_allow_html=True)
            with col_del:
                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
                eid = ev["id"]
                if st.button("🗑", key=f"del_{key_prefix}_{eid}", help="מחק"):
                    if owner and data is not None:
                        data["events"][owner] = [e for e in data["events"][owner] if e["id"] != eid]
                        save_data(data)
                        st.rerun()
        else:
            st.markdown(html, unsafe_allow_html=True)


def _event_form(form_key, data, target_brigade=None, is_global=False):
    """Shared add-event form. target_brigade=None means global."""
    from datetime import date as dt_date, timedelta
    today = dt_date.today()
    with st.form(form_key, clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            ev_title = st.text_input("כותרת האירוע *", placeholder="מה האירוע?")
            ev_date  = st.date_input("תאריך", value=today,
                                     min_value=today, max_value=today + timedelta(days=30))
        with c2:
            ev_time  = st.text_input("שעה", placeholder="לדוג׳ 08:00")
            ev_color = st.selectbox("עדיפות / צבע", list(EVENT_COLORS.keys()))
        ev_desc = st.text_area("פירוט (אופציונלי)", placeholder="פרטים נוספים...", height=70)
        lbl = "שלח לכל החטמ\"רים ★" if is_global else "הוסף לו\"ז"
        if st.form_submit_button(lbl, type="primary", use_container_width=True) and ev_title.strip():
            new_ev = {
                "id":    gen_id(),
                "title": ev_title.strip(),
                "date":  ev_date.isoformat(),
                "time":  ev_time.strip(),
                "desc":  ev_desc.strip(),
                "color": ev_color,
            }
            if is_global:
                data.setdefault("global_events", []).append(new_ev)
            else:
                data["events"].setdefault(target_brigade, []).append(new_ev)
            save_data(data)
            st.rerun()


def show_schedule_tab(brigade, data):
    """Schedule tab for a brigade: shows global events (read-only) + own events."""
    # Merge global + local for the calendar display
    g_events = [dict(ev, is_global=True) for ev in data.get("global_events", [])]
    local_events = data["events"].get(brigade, [])

    all_events = g_events + local_events

    if not all_events:
        st.info("אין עדיין אירועים בלו\"ז. הוסף אירוע למטה.")
    else:
        st.markdown("### 📅 לו\"ז שבועיים קרובים")
        show_calendar(all_events, key_prefix=f"sched_{brigade}",
                      can_delete=True, owner=brigade, data=data)

    st.markdown("---")
    st.markdown("**➕ הוסף אירוע ללו\"ז שלי**")
    _event_form(f"add_ev_{brigade}", data, target_brigade=brigade, is_global=False)


# ─── DIVISION RABBI VIEW ──────────────────────────────────────────────────────
def show_division_view(data):
    # ── Overview grid ──
    st.markdown(f'<div class="section-header"><div class="section-icon">📊</div><p class="section-title">סקירה כללית - כל החטמ"רים</p></div>', unsafe_allow_html=True)
    cols = st.columns(3)
    for i, brigade in enumerate(BRIGADES):
        total, done = completion_stats(data, brigade)
        pct  = round((done / total) * 100) if total else 0
        color = pct_color(pct)
        with cols[i % 3]:
            st.markdown(f"""
            <div class="brigade-card" style="border-top:4px solid {color}">
                <h3 style="margin:0;color:#374151;font-size:1rem">חטמ"ר {brigade}</h3>
                <div class="brigade-pct" style="color:{color}">{pct}%</div>
                <p style="font-size:0.78rem;color:#6b7280;margin:0">{done} / {total} משימות</p>
                <div class="progress-wrap" style="margin-top:8px">
                    <div class="progress-bar" style="width:{pct}%;background:{color}"></div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    st.markdown("---")

    # ── Global tasks management ──
    st.markdown(f'<div class="section-header"><div class="section-icon" style="background:{GOLD};color:#2E3521">📡</div><p class="section-title">משימות אוגדתיות (לכולם)</p></div>', unsafe_allow_html=True)

    if not data["global_tasks"]:
        st.info("אין עדיין משימות אוגדתיות")
    for task in data["global_tasks"]:
        col_t, col_d = st.columns([9, 1])
        with col_t:
            badges = ""
            if task.get("urgent"):   badges += '<span class="badge badge-urgent">⚠ חובת עדכון</span>'
            if task.get("category"): badges += f'<span class="badge badge-cat">{task["category"]}</span>'
            st.markdown(f"""
            <div class="task-card" style="border-right-color:{GOLD}">
                <div class="task-title">{task['title']} {badges}</div>
                {f'<div class="task-desc">{task["desc"]}</div>' if task.get("desc") else ""}
            </div>""", unsafe_allow_html=True)
        with col_d:
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            if st.button("🗑", key=f"del_global_{task['id']}", help="מחק מכולם"):
                data["global_tasks"] = [t for t in data["global_tasks"] if t["id"] != task["id"]]
                for b in BRIGADES:
                    data["completions"].get(b, {}).pop(f"g_{task['id']}", None)
                save_data(data)
                st.rerun()

    st.markdown("---")
    st.markdown("**➕ הוסף משימה לכל החטמ\"רים**")
    with st.form("add_global", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            title  = st.text_input("כותרת *", placeholder="כותרת ברורה")
            cat    = st.selectbox("קטגוריה", ["", "פורים", "פסח", "כללי"])
        with col2:
            desc   = st.text_area("פירוט", placeholder="הוראות נוספות...", height=100)
            urgent = st.checkbox("⚠ חובת עדכון")
        if st.form_submit_button("שלח לכל החטמ\"רים", type="primary", use_container_width=True) and title.strip():
            data["global_tasks"].append({"id": gen_id(), "title": title.strip(), "desc": desc.strip(), "urgent": urgent, "category": cat})
            save_data(data)
            st.rerun()

    st.markdown("---")

    # ── Global events management ──
    st.markdown(f'<div class="section-header"><div class="section-icon" style="background:{GOLD};color:#2E3521">★</div><p class="section-title">אירועים בלו"ז - לכל החטמ"רים</p></div>', unsafe_allow_html=True)

    g_events = data.get("global_events", [])
    if not g_events:
        st.info("אין עדיין אירועים אוגדתיים")
    for ev in g_events:
        col_e, col_d = st.columns([10, 1])
        with col_e:
            highlight_html = f'<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:6px;padding:6px 10px;margin-top:6px;font-size:0.8rem;color:#991b1b">🔴 <strong>{ev["highlight"]}</strong></div>' if ev.get("highlight") else ""
            st.markdown(f'''
            <div style="background:white;border-radius:10px;padding:14px;border-right:4px solid {GOLD};
                        box-shadow:0 2px 6px rgba(0,0,0,0.06);margin-bottom:8px">
                <span style="background:#f3f4f6;padding:2px 8px;border-radius:999px;font-size:0.76rem;font-weight:700">יום {ev.get('day','')}</span>
                {f"<span style='margin-right:8px;font-size:0.8rem;color:#6b7280'>🕐 {ev['time']}</span>" if ev.get('time') else ""}
                <strong style='margin-right:8px'>{ev['title']}</strong>
                {f"<span style='font-size:0.82rem;color:#4b5563'>{ev['desc']}</span>" if ev.get('desc') else ""}
                {highlight_html}
            </div>''', unsafe_allow_html=True)
        with col_d:
            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
            if st.button("🗑", key=f"del_gev_{ev['id']}", help="מחק מכללם"):
                data["global_events"] = [e for e in data["global_events"] if e["id"] != ev["id"]]
                save_data(data)
                st.rerun()

    st.markdown("**➕ הוסף אירוע ללו\"ז לכלל החטמ\"רים**")
    with st.form("add_global_event", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            gev_title = st.text_input("כותרת האירוע *", placeholder="מה האירוע?")
            gev_day   = st.selectbox("יום", DAYS, key="gev_day")
        with c2:
            gev_time  = st.text_input("שעה", placeholder="לדוג׳ 09:00", key="gev_time")
            gev_desc  = st.text_area("פירוט", placeholder="הוראות...", height=70, key="gev_desc")
        gev_highlight = st.text_input("📌 הדגשה (השאר ריק אם לא נחוץ)", placeholder="לדוג׳: נוכחות חובה!", key="gev_highlight")
        if st.form_submit_button("שלח לכל החטמ\"רים ★", type="primary", use_container_width=True) and gev_title.strip():
            data["global_events"].append({
                "id": gen_id(), "title": gev_title.strip(), "day": gev_day,
                "time": gev_time.strip(), "desc": gev_desc.strip(),
                "highlight": gev_highlight.strip()
            })
            save_data(data)
            st.rerun()


# ─── EXCEL EXPORT ────────────────────────────────────────────────────────────
def generate_excel(data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill  = PatternFill("solid", fgColor="4E5A37")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    title_fill   = PatternFill("solid", fgColor="D4AF37")
    title_font   = Font(bold=True, color="2E3521", name="Arial", size=13)
    done_fill    = PatternFill("solid", fgColor="D1FAE5")
    undone_fill  = PatternFill("solid", fgColor="FEF2F2")
    thin = Border(left=Side(style='thin', color='D1D5DB'), right=Side(style='thin', color='D1D5DB'),
                  top=Side(style='thin', color='D1D5DB'),  bottom=Side(style='thin', color='D1D5DB'))
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center', wrap_text=True)

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = center; cell.border = thin

    # ── Overview sheet ──
    ws = wb.create_sheet("סקירה כללית")
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.merge_cells("A1:D1")
    ws["A1"] = 'סקירה כללית - מערך הרבנות'
    ws["A1"].fill = title_fill; ws["A1"].font = title_font
    ws["A1"].alignment = center; ws["A1"].border = thin
    for col, header in enumerate(['חטמ"ר', 'סה"כ משימות', 'הושלמו', 'אחוז ביצוע'], start=1):
        ws.cell(row=2, column=col, value=header)
    style_header(ws, 2, 4)
    for r, brigade in enumerate(BRIGADES, start=3):
        total, done = completion_stats(data, brigade)
        pct = round((done / total) * 100) if total else 0
        for col, val in enumerate([brigade, total, done, f"{pct}%"], start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = center; cell.border = thin
            cell.fill = done_fill if pct == 100 else PatternFill("solid", fgColor="FEF9C3") if pct >= 50 else undone_fill
        ws.row_dimensions[r].height = 20

    # ── Per brigade sheet ──
    for brigade in BRIGADES:
        ws = wb.create_sheet(brigade)
        ws.sheet_view.rightToLeft = True
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.merge_cells("A1:E1")
        ws["A1"] = f'חטמ"ר {brigade} - משימות פורים ופסח'
        ws["A1"].fill = title_fill; ws["A1"].font = title_font
        ws["A1"].alignment = center
        for col, h in enumerate(['כותרת', 'פירוט', 'קטגוריה', 'חובת עדכון', 'בוצע'], start=1):
            ws.cell(row=2, column=col, value=h)
        style_header(ws, 2, 5)
        comp = data["completions"].get(brigade, {})
        row = 3
        for task in data["global_tasks"]:
            ck = f"g_{task['id']}"
            done = bool(comp.get(ck, False))
            vals = [task['title'], task.get('desc', ''), task.get('category', ''), 'כן' if task.get('urgent') else '', '✓' if done else '']
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = right if col <= 2 else center
                cell.border = thin
                if done: cell.fill = done_fill
            row += 1

        for task in data["brigade_tasks"].get(brigade, []):
            ck = f"l_{task['id']}"
            done = bool(comp.get(ck, False))
            vals = [task['title'], task.get('desc', ''), '', 'כן' if task.get('urgent') else '', '✓' if done else '']
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = right if col <= 2 else center
                cell.border = thin
                if done: cell.fill = done_fill
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    inject_css()
    data = load_data()

    # ── Header ──
    st.markdown(f"""
    <div class="app-header">
        <div style="font-size:2.2rem">✡️</div>
        <div>
            <h1>מערך הרבנות - פקודות ומשימות</h1>
            <p>סדר שבועי: היערכות לפורים ופסח</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Excel export (always visible) ──
    col_xl, _ = st.columns([2, 8])
    with col_xl:
        if st.button("📥 ייצוא לאקסל", key="btn_excel"):
            try:
                excel_buf = generate_excel(data)
                st.download_button(
                    label="📥 לחץ להורדה",
                    data=excel_buf,
                    file_name=f"משימות_רבנות_{datetime.now().strftime('%d%m%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_excel"
                )
            except Exception as e:
                st.error(f"שגיאה בייצוא: {e}")

    st.markdown("<div style='margin-top:8px'></div>", unsafe_allow_html=True)

    # ── Browser-style tabs: one per role ──
    tab_labels = ["👑 רב האוגדה"] + [f"🪖 {b}" for b in BRIGADES]
    tabs = st.tabs(tab_labels)

    with tabs[0]:
        # ── Password protection for division rabbi ──
        if "division_unlocked" not in st.session_state:
            st.session_state.division_unlocked = False
        if not st.session_state.division_unlocked:
            st.markdown(f"""
            <div style="background:white;border-radius:16px;padding:40px;max-width:400px;margin:60px auto;
                        box-shadow:0 4px 24px rgba(0,0,0,0.12);text-align:center">
                <div style="font-size:3rem;margin-bottom:16px">🔐</div>
                <h2 style="color:{IDF_GREEN};margin-bottom:8px">טאב רב האוגדה</h2>
                <p style="color:#6b7280;font-size:0.9rem">הכנס סיסמא כדי להמשיך</p>
            </div>
            """, unsafe_allow_html=True)
            col_pw, _ = st.columns([2, 4])
            with col_pw:
                pw = st.text_input("סיסמא", type="password", key="pw_input",
                                   placeholder="הכנס סיסמא...", label_visibility="collapsed")
                if st.button("פתח 🔓", use_container_width=True, type="primary"):
                    if pw == "12345":
                        st.session_state.division_unlocked = True
                        st.rerun()
                    else:
                        st.error("סיסמא שגויה ❌")
        else:
            if st.button("🔒 נעל", key="lock_division", help="נעל את הטאב"):
                st.session_state.division_unlocked = False
                st.rerun()
            show_division_view(data)

    for i, brigade in enumerate(BRIGADES):
        with tabs[i + 1]:
            # Mini header inside each brigade tab
            total, done = completion_stats(data, brigade)
            pct = round((done / total) * 100) if total else 0
            color = pct_color(pct)
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:16px;background:white;border-radius:12px;
                        padding:12px 20px;margin-bottom:16px;box-shadow:0 2px 8px rgba(0,0,0,0.06)">
                <div style="font-size:1.8rem;font-weight:800;color:{color}">{pct}%</div>
                <div>
                    <div style="font-weight:800;font-size:1.1rem;color:#1f2937">חטמ"ר {brigade}</div>
                    <div style="font-size:0.8rem;color:#6b7280">{done} מתוך {total} משימות הושלמו</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            show_brigade_view(brigade, data)

if __name__ == "__main__":
    main()
